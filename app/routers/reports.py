from datetime import date, timedelta
from io import BytesIO
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from app.db import get_db
from app.services.payroll import PayrollService
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

router = APIRouter(
    prefix="/reports",
    tags=["reports"]
)

@router.get("/export/payroll")
async def export_payroll(
    start_date: date = Query(None),
    end_date: date = Query(None),
    branch_id: int = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """
    Generate and download Global Payroll Excel Report.
    """
    if not start_date:
        today = date.today()
        start_date = today.replace(day=1)
    if not end_date:
        today = date.today()
        # End of current month approx logic or just today
        if today.month == 12:
            end_date = date(today.year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(today.year, today.month + 1, 1) - timedelta(days=1)

    # 1. Get Data
    payroll_data = await PayrollService.get_payroll_stats(db, start_date, end_date, branch_id)
    
    # 2. Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "État de Paie (Détails)"
    
    # Headers - Manual Worksheet Style
    headers = [
        "Store", "Employé", "CIN", "Fonction", "Sal. Base", 
        "Absences (Jours)", "Détail Absences", 
        "Congés (Jours Non Payés)", "Détail Congés",
        "Jours Déduits Total", "Déduction (Est.)", # NEW
        "Total Avances", "Détail Avances",
        "Prêts (Dû)", 
        "Ventes (Qty)", "Ventes (Rev)", 
        "Net à Payer (Est.)", "Notes / Signature"
    ]
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid") # Green for "Worksheet"
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Draw Headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Draw Rows
    row_num = 2
    for item in payroll_data:
        emp = item["employee"]
        stats = item["stats"]
        
        branch_name = emp.branch.name if emp.branch else "N/A"
        full_name = f"{emp.first_name} {emp.last_name}"
        
        # --- Format Details ---
        
        # Absences: "01/01 \n 05/01"
        abs_dates = [a.date.strftime('%d/%m') for a in stats["absences_list"]]
        abs_str = "\n".join(abs_dates) if abs_dates else "-"
        
        # Advances: "01/01: 50dt \n 10/01: 20dt"
        adv_details = [f"{d.date.strftime('%d/%m')}: {d.amount:.0f}" for d in stats["advances_list"]]
        adv_str = "\n".join(adv_details) if adv_details else "-"
        
        # Leaves: "01/01->05/01 (Payé)"
        leave_details = []
        for l in stats["leaves_list"]:
            l_type_map = {
                "paid": "Payé",
                "unpaid": "Non Payé",
                "sick": "Maladie (Payé)",
                "sick_unpaid": "Maladie (Non Payé)"
            }
            l_type = l_type_map.get(l.ltype.value, l.ltype.value)
            leave_details.append(f"{l.start_date.strftime('%d/%m')}->{l.end_date.strftime('%d/%m')} ({l_type})")
        leave_str = "\n".join(leave_details) if leave_details else "-"

        row = [
            branch_name,
            full_name,
            emp.cin or "",
            emp.position,
            float(stats["salary"]),
            stats["absences_count"],
            abs_str,
            stats.get("unpaid_leave_days", 0), # New: Computed Unpaid Days
            leave_str,
            stats.get("total_deduction_days", 0), # New: Total Days Deducted
            float(stats.get("deduction_amount", 0)), # New: Deduction Amount
            float(stats["advances_total"]),
            adv_str,
            float(stats["loans_due_total"]),
            stats["sales_qty"],
            float(stats["sales_rev"]),
            float(stats.get("net_estimated", 0)), # New: Net
            "" # Signature
        ]
        
        for col_num, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.border = thin_border
            # Wrap text for detail columns
            if col_num in [7, 9, 13]: 
                cell.alignment = left_align
            else:
                cell.alignment = center_align
                
            # Format Currency columns: Sal(5), Ded(11), Av(12), Prêt(14), Rev(16), Net(17)
            if col_num in [5, 11, 12, 14, 16, 17]:
                cell.number_format = '#,##0.000 "DT"'
            
        row_num += 1

    # Adjust Column Widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['G'].width = 15 # Abs Detail
    ws.column_dimensions['I'].width = 25 # Leaves Detail
    ws.column_dimensions['L'].width = 15 # Av Total
    ws.column_dimensions['Q'].width = 18 # Net
    ws.column_dimensions['R'].width = 20 # Sig

    # 3. Save to Stream
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Payroll_Worksheet_{start_date}_{end_date}.xlsx"
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
